from selenium import webdriver
import time
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
nos=0
with open("input.txt", "r") as a_file:
    for line in a_file:
        item = line.strip()
        nos+=1
        ws1 = wb.create_sheet(str(nos)+"Alibaba")
        ws2 = wb.create_sheet(str(nos)+"Indiamart")
        row1 = ws1.row_dimensions[1]
        row2 = ws2.row_dimensions[1]
        row1.font = Font(bold=True)
        row2.font = Font(bold=True)
        ws1['A1']='product_Name'
        ws1['B1']='product_Price'
        ws1['C1']='product_Quantity'
        ws1['D1']='seller_Name'
        ws1['E1']='seller_Rating'
        ws1['F1']='number_Of_Ratings'

        ws2['A1']='product_Name'
        ws2['B1']='product_Price'
        ws2['C1']='seller_Name'

        class BoT:
            def __init__(self):
                self.driver = webdriver.Firefox()

            def closeBrowser(self):
                self.driver.close()

            def DoIt(self):
                drive = self.driver
                products_Page=drive.get("https://www.alibaba.com/trade/search?fsb=y&IndexArea=product_en&CatId=&SearchText=" + item)
                time.sleep(1)
                temp=2

                for j in range(2):
                    drive.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                    time.sleep(3)
                    products_List=drive.find_elements_by_xpath("//*[@class='organic-gallery-title__content medium']")
                    products_Price=drive.find_elements_by_xpath("//*[@class='gallery-offer-price medium']")
                    products_Quantity=drive.find_elements_by_xpath("//*[@class='gallery-offer-minorder medium']")
                    seller_Name=drive.find_elements_by_xpath("//*[@class='organic-list-offer__seller-company']")
                    seller_Rating=drive.find_elements_by_xpath("//*[@class='seb-supplier-review__score']")
                    number_Of_Ratings=drive.find_elements_by_xpath("//*[@class='seb-supplier-review__review-count']")


                    k=temp
                    for i in products_List:
                        ws1.cell(k, 1, i.text)
                        k+=1
                    k=temp
                    for i in products_Quantity:
                        ws1.cell(k, 3, i.text)
                        k+=1
                    k=temp
                    for i in seller_Name:
                        ws1.cell(k, 4,i.text)
                        k+=1
                    k=temp
                    for i in seller_Rating:
                        ws1.cell(k,5 , i.text)
                        k+=1
                    k=temp
                    for i in number_Of_Ratings:
                        ws1.cell(k,6 , i.text)
                        k+=1
                    k=temp
                    for i in products_Price:
                        ws1.cell(k, 2, i.text)
                        k += 1
                        temp=k

                    nxt = drive.get("https://www.alibaba.com/products/" + item + ".html?spm=a2700.galleryofferlist.0.0.79812f4034Fl8f&IndexArea=product_en&page=" + str(j + 1))
                    print(temp)


        class bOt:
            def __init__(self):
                self.driver = webdriver.Firefox()

            def closeBrowser(self):
                self.driver.close()

            def LeGo(self):
                drive = self.driver
                time.sleep(1)
                products_Page=drive.get("https://dir.indiamart.com/search.mp?ss= "+ item)
                time.sleep(2)
                temp1=2

                for j in range(1):
                    drive.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                    time.sleep(2)
                    drive.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                    time.sleep(2)
                    drive.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                    time.sleep(2)
                    products_List1=drive.find_elements_by_xpath("//*[@class='lg']")
                    products_Price1=drive.find_elements_by_xpath("//*[@class='prc cur']")
                    seller_Name1=drive.find_elements_by_xpath("//*[@class='lcname']")

                    k=temp1
                    for i in products_List1:
                        ws2.cell(k, 1, i.text)
                        k+=1
                    k=temp1
                    for i in seller_Name1:
                        ws2.cell(k, 3,i.text)
                        k+=1
                    k=temp1
                    for i in products_Price1:
                        ws2.cell(k, 2, i.text)
                        k += 1
                        temp1=k

                    print(temp1)


        BoT().DoIt()
        time.sleep(3)
        bOt().LeGo()
wb.save("Product Listings.xlsx")

